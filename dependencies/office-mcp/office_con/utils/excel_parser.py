import io

import openpyxl
import xlrd


class ExcelParser:
    """Parses .xls and .xlsx files into a 2D string grid (values)."""

    def __init__(self, data: bytes, filename: str, remove_blank: bool = False) -> None:
        self.values: list[list[str]] = [[]]
        if filename.lower().endswith(".xls"):
            self.parse_xls(data, filename)
        elif filename.lower().endswith(".xlsx"):
            self.parse_xlsx(data, filename)
        else:
            raise ValueError(f"Unsupported file format {filename}")
        # remove completely blank rows and columns
        if remove_blank:
            self.remove_empty_rows()
            self.remove_empty_columns()
        # normalize all cells to string type
        for row_index in range(len(self.values)):
            self.values[row_index] = [str(cell if cell is not None else "").strip() for cell in self.values[row_index]]

    def remove_empty_rows(self) -> None:
        for row_index in reversed(range(len(self.values))):
            row = self.values[row_index]
            if all(cell == "" for cell in row):
                del self.values[row_index]

    def remove_empty_columns(self) -> None:
        col_count = len(self.values[0])
        for i in reversed(range(col_count)):
            if all(row[i] == "" for row in self.values):
                # remove column
                for row in self.values:
                    del row[i]

    def parse_xlsx(self, data: bytes, filename: str) -> None:
        workbook = openpyxl.load_workbook(filename=io.BytesIO(data))
        sheet_names = workbook.sheetnames
        worksheet = workbook[sheet_names[0]]
        for row_index, row in enumerate(worksheet.iter_rows()):
            # skip zero height rows
            hidden = worksheet.row_dimensions[row_index].hidden
            if hidden:
                continue
            self.values.append([str(cell.value) if cell.value is not None else "" for cell in row])

    def parse_xls(self, data: bytes, filename: str) -> None:
        wb = xlrd.open_workbook(file_contents=data)
        # iterate through sheets
        ws = wb.sheet_by_index(0)
        # terminate the last valid row and column
        last_row = -1
        last_col = -1
        for row_index, row in enumerate(ws.get_rows()):
            for col_index, cell in enumerate(row):
                if cell.value is None:
                    break
                last_row = max(last_row, row_index)
                last_col = max(last_col, col_index)

        # create evenly siyzed array of strings
        self.values = [["" for _ in range(last_col + 1)] for _ in range(last_row + 1)]

        # fill raw data with values
        for row_index, row in enumerate(ws.get_rows()):
            for col_index, cell in enumerate(row):
                self.values[row_index][col_index] = cell.value
